import csv
import io
import logging
from pathlib import Path

logger = logging.getLogger(__name__)


def extract_text(file_path: Path, ocr_enabled: bool = True) -> str | None:
    suffix = file_path.suffix.lower()
    try:
        if suffix in (".txt", ".md"):
            return _extract_plaintext(file_path)
        elif suffix == ".pdf":
            return _extract_pdf(file_path)
        elif suffix == ".docx":
            return _extract_docx(file_path)
        elif suffix == ".xlsx":
            return _extract_xlsx(file_path)
        elif suffix == ".csv":
            return _extract_csv(file_path)
        elif suffix in (".png", ".jpg", ".jpeg", ".tiff", ".bmp"):
            if ocr_enabled:
                return _extract_image_ocr(file_path)
            else:
                logger.info(f"OCR disabled, skipping image: {file_path}")
                return None
        else:
            logger.warning(f"Unsupported file type: {suffix} for {file_path}")
            return None
    except Exception as e:
        logger.error(f"Failed to extract text from {file_path}: {e}")
        return None


def _extract_plaintext(file_path: Path) -> str:
    return file_path.read_text(encoding="utf-8", errors="replace")


def _extract_pdf(file_path: Path) -> str:
    import fitz  # pymupdf

    text_parts = []
    with fitz.open(str(file_path)) as doc:
        for page in doc:
            text_parts.append(page.get_text())
    return "\n".join(text_parts)


def _extract_docx(file_path: Path) -> str:
    from docx import Document

    doc = Document(str(file_path))
    paragraphs = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n".join(paragraphs)


def _extract_xlsx(file_path: Path) -> str:
    from openpyxl import load_workbook

    wb = load_workbook(str(file_path), read_only=True, data_only=True)
    text_parts = []
    for sheet in wb.worksheets:
        text_parts.append(f"[Sheet: {sheet.title}]")
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) if c is not None else "" for c in row]
            if any(cells):
                text_parts.append(" | ".join(cells))
    wb.close()
    return "\n".join(text_parts)


def _extract_csv(file_path: Path) -> str:
    text = file_path.read_text(encoding="utf-8", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = [" | ".join(row) for row in reader if any(row)]
    return "\n".join(rows)


def _extract_image_ocr(file_path: Path) -> str | None:
    try:
        import pytesseract
        from PIL import Image

        img = Image.open(file_path)
        text = pytesseract.image_to_string(img)
        return text.strip() if text.strip() else None
    except ImportError:
        logger.warning("pytesseract or Pillow not installed, skipping OCR")
        return None
    except Exception as e:
        logger.error(f"OCR failed for {file_path}: {e}")
        return None
